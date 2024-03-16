import sys
from pinyin import pinyin
import openpyxl


def pinyin(hanzi):
     '''
     将汉字变为pinyin
     format choice [strip,numerical]
     '''
     return pinyin.get(hanzi,format="strip",delimiter="")

def reading_in_excel(excel):
    """
    """
    uniq_list = []
    book = openpyxl.load_workbook(excel)
    #print(book.sheetnames)
    sheet3 = book['Sheet3']
    pinyin_sheet = book.create_sheet(title='exchange_to_pinyin')
    for row in sheet3:
        row_list = [cell.value for cell in row]
        name = row_list[0]
        novoid = row_list[2]
        pinyin_name = pinyin(name)
        if pinyin_name not in uniq_list:
            uniq_list.append(pinyin_name)
        else:
            print('Warning! {pinyin_name} appeared twice!'.format(pinyin_name=pinyin_name))
        new_list = [name,pinyin_name,novoid]
        pinyin_sheet.append(new_list)
    book.save('exchanged.xlsx')

if __name__ == "__main__":
    excel = sys.argv[1]
    reading_in_excel(excel)